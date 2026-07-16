#!/usr/bin/env python3
"""
Automated tests for validate_custom_instructions.py

Tests cover:
- No custom instruction file found (existence validation failure)
- Valid file within the 4000-character limit
- Invalid file above the 4000-character limit (size validation failure)

Usage:
    python test_validate_custom_instructions.py
    python -m pytest test_validate_custom_instructions.py -v
"""

import base64
import json
import sys
import unittest
from unittest.mock import MagicMock, patch, call

# Ensure the module under test is importable from this directory
sys.path.insert(0, '.')

import validate_custom_instructions as vci


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _make_file_api_response(content: str) -> dict:
    """Simulate a GitHub API file response with base64-encoded content."""
    encoded = base64.b64encode(content.encode('utf-8')).decode('utf-8')
    return {
        'type': 'file',
        'content': encoded,
        'encoding': 'base64',
    }


# ---------------------------------------------------------------------------
# Unit tests: decode_file_content
# ---------------------------------------------------------------------------

class TestDecodeFileContent(unittest.TestCase):
    """Tests for the decode_file_content helper."""

    def test_decodes_valid_base64_content(self):
        text = "Hello, Copilot!"
        response = _make_file_api_response(text)
        result = vci.decode_file_content(response)
        self.assertEqual(result, text)

    def test_returns_none_for_non_dict_response(self):
        self.assertIsNone(vci.decode_file_content(None))
        self.assertIsNone(vci.decode_file_content([]))

    def test_returns_none_for_empty_content_field(self):
        self.assertIsNone(vci.decode_file_content({'content': ''}))

    def test_returns_none_for_missing_content_field(self):
        self.assertIsNone(vci.decode_file_content({'type': 'file'}))


# ---------------------------------------------------------------------------
# Unit tests: validate_existence
# ---------------------------------------------------------------------------

class TestValidateExistence(unittest.TestCase):
    """Tests for the validate_existence validation function."""

    def test_fails_when_no_files_found(self):
        """Scenario: No custom instruction file found in any repository."""
        results = [
            {'repo': 'org/repo1', 'files_found': [], 'violations': [], 'has_custom_instructions': False, 'error': None},
            {'repo': 'org/repo2', 'files_found': [], 'violations': [], 'has_custom_instructions': False, 'error': None},
        ]
        self.assertFalse(vci.validate_existence(results))

    def test_passes_when_at_least_one_file_found(self):
        """Scenario: At least one custom instruction file exists."""
        results = [
            {
                'repo': 'org/repo1',
                'files_found': [
                    {'path': '.github/copilot-instructions.md', 'char_count': 100, 'exceeds_limit': False}
                ],
                'violations': [],
                'has_custom_instructions': True,
                'error': None,
            },
        ]
        self.assertTrue(vci.validate_existence(results))

    def test_fails_with_empty_results_list(self):
        """Edge case: Empty results (no repos scanned)."""
        self.assertFalse(vci.validate_existence([]))


# ---------------------------------------------------------------------------
# Unit tests: validate_file_sizes
# ---------------------------------------------------------------------------

class TestValidateFileSizes(unittest.TestCase):
    """Tests for the validate_file_sizes validation function."""

    def test_returns_empty_list_when_all_files_within_limit(self):
        """Scenario: Valid file within the 4000-character limit."""
        results = [
            {
                'repo': 'org/repo1',
                'files_found': [
                    {'path': '.github/copilot-instructions.md', 'char_count': 500, 'exceeds_limit': False}
                ],
                'violations': [],
                'has_custom_instructions': True,
                'error': None,
            },
        ]
        violations = vci.validate_file_sizes(results)
        self.assertEqual(violations, [])

    def test_returns_violations_when_file_exceeds_limit(self):
        """Scenario: Invalid file above 4000 characters."""
        violation = {
            'repository': 'org/repo1',
            'file_path': '.github/copilot-instructions.md',
            'char_count': 4500,
            'limit': 4000,
            'excess_chars': 500,
        }
        results = [
            {
                'repo': 'org/repo1',
                'files_found': [
                    {'path': '.github/copilot-instructions.md', 'char_count': 4500, 'exceeds_limit': True}
                ],
                'violations': [violation],
                'has_custom_instructions': True,
                'error': None,
            },
        ]
        violations = vci.validate_file_sizes(results)
        self.assertEqual(len(violations), 1)
        self.assertEqual(violations[0]['char_count'], 4500)
        self.assertEqual(violations[0]['excess_chars'], 500)

    def test_returns_violations_for_multiple_repos(self):
        """Multiple repositories with multiple violations."""
        v1 = {
            'repository': 'org/repo1',
            'file_path': '.github/copilot-instructions.md',
            'char_count': 5000,
            'limit': 4000,
            'excess_chars': 1000,
        }
        v2 = {
            'repository': 'org/repo2',
            'file_path': '.github/instructions/coding.instructions.md',
            'char_count': 4001,
            'limit': 4000,
            'excess_chars': 1,
        }
        results = [
            {'repo': 'org/repo1', 'files_found': [], 'violations': [v1], 'has_custom_instructions': True, 'error': None},
            {'repo': 'org/repo2', 'files_found': [], 'violations': [v2], 'has_custom_instructions': True, 'error': None},
        ]
        violations = vci.validate_file_sizes(results)
        self.assertEqual(len(violations), 2)

    def test_file_at_exact_limit_is_valid(self):
        """A file with exactly 4000 characters should not be flagged."""
        results = [
            {
                'repo': 'org/repo1',
                'files_found': [
                    {'path': '.github/copilot-instructions.md', 'char_count': 4000, 'exceeds_limit': False}
                ],
                'violations': [],
                'has_custom_instructions': True,
                'error': None,
            },
        ]
        violations = vci.validate_file_sizes(results)
        self.assertEqual(violations, [])


# ---------------------------------------------------------------------------
# Unit tests: assess_repo (integration of get_custom_instruction_files)
# ---------------------------------------------------------------------------

class TestAssessRepo(unittest.TestCase):
    """Tests for assess_repo function using mocked GitHub CLI calls."""

    def _mock_repo(self, name='org/repo'):
        return {'nameWithOwner': name}

    @patch('validate_custom_instructions.run_gh_command')
    def test_no_custom_instruction_files_found(self, mock_run):
        """Scenario: Repository has no custom instruction files."""
        # All API calls return None (file not found)
        mock_run.return_value = None

        result = vci.assess_repo(self._mock_repo())

        self.assertFalse(result['has_custom_instructions'])
        self.assertEqual(result['files_found'], [])
        self.assertEqual(result['violations'], [])
        self.assertIsNone(result['error'])

    @patch('validate_custom_instructions.run_gh_command')
    def test_valid_file_within_limit(self, mock_run):
        """Scenario: Repository has a valid custom instruction file within the 4000-char limit."""
        valid_content = "Follow PEP 8 coding standards.\n" * 10  # ~310 chars
        file_response = _make_file_api_response(valid_content)

        def side_effect(command):
            if 'copilot-instructions.md' in command and '/contents/' in command and 'instructions/' not in command:
                return file_response
            return None  # instructions/ dir not found

        mock_run.side_effect = side_effect

        result = vci.assess_repo(self._mock_repo())

        self.assertTrue(result['has_custom_instructions'])
        self.assertEqual(len(result['files_found']), 1)
        self.assertEqual(result['violations'], [])
        self.assertFalse(result['files_found'][0]['exceeds_limit'])

    @patch('validate_custom_instructions.run_gh_command')
    def test_invalid_file_above_limit(self, mock_run):
        """Scenario: Repository has a custom instruction file exceeding 4000 characters."""
        over_limit_content = "A" * 4500  # 4500 chars, over the 4000 limit
        file_response = _make_file_api_response(over_limit_content)

        def side_effect(command):
            if 'copilot-instructions.md' in command and '/contents/' in command and 'instructions/' not in command:
                return file_response
            return None

        mock_run.side_effect = side_effect

        result = vci.assess_repo(self._mock_repo())

        self.assertTrue(result['has_custom_instructions'])
        self.assertEqual(len(result['violations']), 1)
        violation = result['violations'][0]
        self.assertEqual(violation['char_count'], 4500)
        self.assertEqual(violation['limit'], 4000)
        self.assertEqual(violation['excess_chars'], 500)

    @patch('validate_custom_instructions.run_gh_command')
    def test_instructions_dir_files_are_checked(self, mock_run):
        """
        Scenario: .github/instructions/ contains an over-limit .instructions.md file.
        """
        over_limit_content = "B" * 4001
        file_in_dir_response = _make_file_api_response(over_limit_content)

        dir_listing = [
            {
                'type': 'file',
                'name': 'coding.instructions.md',
                'path': '.github/instructions/coding.instructions.md',
            }
        ]

        def side_effect(command):
            if 'copilot-instructions.md' in command and 'instructions/' not in command:
                return None  # root copilot-instructions.md not present
            if 'contents/.github/instructions"' in command or command.endswith('contents/.github/instructions'):
                return dir_listing
            if 'coding.instructions.md' in command:
                return file_in_dir_response
            return None

        mock_run.side_effect = side_effect

        result = vci.assess_repo(self._mock_repo())

        self.assertTrue(result['has_custom_instructions'])
        self.assertEqual(len(result['violations']), 1)
        self.assertEqual(result['violations'][0]['char_count'], 4001)

    @patch('validate_custom_instructions.run_gh_command')
    def test_file_at_exact_limit_is_not_a_violation(self, mock_run):
        """A file with exactly 4000 characters must not be a violation."""
        exact_content = "C" * 4000
        file_response = _make_file_api_response(exact_content)

        def side_effect(command):
            if 'copilot-instructions.md' in command and 'instructions/' not in command:
                return file_response
            return None

        mock_run.side_effect = side_effect

        result = vci.assess_repo(self._mock_repo())

        self.assertTrue(result['has_custom_instructions'])
        self.assertEqual(result['violations'], [])
        self.assertFalse(result['files_found'][0]['exceeds_limit'])


# ---------------------------------------------------------------------------
# Unit tests: export_violations_to_excel
# ---------------------------------------------------------------------------

class TestExportViolationsToExcel(unittest.TestCase):
    """Tests for the Excel export function."""

    def setUp(self):
        import tempfile
        self._tmp = tempfile.mkdtemp()
        vci.CONFIG['output_dir'] = self._tmp

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmp, ignore_errors=True)
        vci.CONFIG['output_dir'] = '.'

    def test_returns_none_when_no_violations(self):
        result = vci.export_violations_to_excel([])
        self.assertIsNone(result)

    def test_creates_excel_file_for_violations(self):
        violations = [
            {
                'repository': 'org/repo1',
                'file_path': '.github/copilot-instructions.md',
                'char_count': 5000,
                'limit': 4000,
                'excess_chars': 1000,
            }
        ]
        output_path = vci.export_violations_to_excel(violations)
        self.assertIsNotNone(output_path)
        import os
        self.assertTrue(os.path.exists(output_path))

    def test_excel_file_contains_correct_data(self):
        violations = [
            {
                'repository': 'org/my-repo',
                'file_path': '.github/copilot-instructions.md',
                'char_count': 4500,
                'limit': 4000,
                'excess_chars': 500,
            }
        ]
        output_path = vci.export_violations_to_excel(violations)
        self.assertIsNotNone(output_path)

        if vci.OPENPYXL_AVAILABLE:
            import openpyxl as ox
            wb = ox.load_workbook(output_path)
            ws = wb.active
            # Row 1 is the header, row 2 is the first data row
            self.assertEqual(ws.cell(row=2, column=1).value, 'org/my-repo')
            self.assertEqual(ws.cell(row=2, column=3).value, 4500)
            self.assertEqual(ws.cell(row=2, column=5).value, 500)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    unittest.main(verbosity=2)
