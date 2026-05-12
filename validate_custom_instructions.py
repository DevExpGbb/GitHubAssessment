#!/usr/bin/env python3
"""
GitHub Copilot Custom Instructions Validation Tool
Validates the existence and size of Repository Custom Instructions files across GitHub repositories.

According to GitHub Copilot documentation, Repository Custom Instructions must not exceed 4000
characters. See: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/
response-customization?tool=webui#about-repository-custom-instructions

Requirements:
    - GitHub CLI (gh) installed and authenticated
    - Python 3.8+
    - openpyxl (optional, for Excel report generation: pip install openpyxl)
      Falls back to CSV output if not installed.

Usage:
    python validate_custom_instructions.py

Configuration:
    Edit the CONFIG section below to customize behavior
"""

import subprocess
import json
import sys
import os
import base64
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from time import time, sleep
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================================
# CONFIGURATION
# ============================================================================

CONFIG = {
    # GitHub CLI command (change if gh is not in PATH)
    'gh_command': 'gh',

    # Custom instruction file patterns to check
    'custom_instruction_files': ['.github/copilot-instructions.md'],
    'custom_instruction_dirs': ['.github/instructions'],
    'custom_instruction_extension': '.instructions.md',

    # Maximum allowed characters per custom instruction file
    # Reference: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/
    # response-customization?tool=webui#about-repository-custom-instructions
    'max_chars': 4000,

    # Performance settings
    'max_workers_fetch': 10,
    'max_workers_check': 15,

    # Rate limiting
    'enable_rate_limit_check': True,
    'rate_limit_threshold': 100,
    'rate_limit_wait_time': 60,
    'request_delay': 0.05,

    # Output settings
    'output_dir': '.',
    'excel_prefix': 'custom_instructions_violations',
    'include_timestamp': True,

    # Personal account identifier (auto-detected from GitHub CLI if empty)
    'personal_account': '',

    # Verbose output
    'verbose': True,
}

# ============================================================================
# CORE FUNCTIONS
# ============================================================================

rate_limit_lock = threading.Lock()
rate_limit_info = {'remaining': None, 'reset_time': None, 'checked': False}


def log(message, verbose_only=False):
    """Print message if verbose or not verbose_only"""
    if not verbose_only or CONFIG['verbose']:
        print(message)


def check_gh_installed():
    """Check if GitHub CLI is installed and accessible"""
    try:
        subprocess.run(
            [CONFIG['gh_command'], '--version'],
            capture_output=True,
            check=True,
            timeout=5
        )
        return True
    except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
        return False


def check_rate_limit():
    """Check GitHub API rate limit status"""
    try:
        result = subprocess.run(
            f"{CONFIG['gh_command']} api rate_limit",
            shell=True,
            capture_output=True,
            text=True,
            check=True,
            timeout=10
        )
        data = json.loads(result.stdout)
        core_rate = data.get('resources', {}).get('core', {})
        return {
            'remaining': core_rate.get('remaining', 5000),
            'limit': core_rate.get('limit', 5000),
            'reset_time': core_rate.get('reset', 0)
        }
    except Exception:
        return None


def wait_for_rate_limit():
    """Wait if rate limit is approaching threshold"""
    if not CONFIG['enable_rate_limit_check']:
        return

    with rate_limit_lock:
        if not rate_limit_info['checked'] or rate_limit_info['remaining'] is None:
            limit_data = check_rate_limit()
            if limit_data:
                rate_limit_info['remaining'] = limit_data['remaining']
                rate_limit_info['reset_time'] = limit_data['reset_time']
                rate_limit_info['checked'] = True

                log(f"📊 Rate Limit: {limit_data['remaining']}/{limit_data['limit']} requests remaining", verbose_only=True)

                if limit_data['remaining'] < CONFIG['rate_limit_threshold']:
                    wait_time = CONFIG['rate_limit_wait_time']
                    log(f"⚠️  Rate limit threshold reached ({limit_data['remaining']} remaining)")
                    log(f"   Waiting {wait_time} seconds before continuing...")
                    sleep(wait_time)
                    rate_limit_info['checked'] = False

        if CONFIG['request_delay'] > 0:
            sleep(CONFIG['request_delay'])


def run_gh_command(command):
    """Run GitHub CLI command and return JSON output"""
    wait_for_rate_limit()

    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            check=True,
            timeout=30
        )
        return json.loads(result.stdout)
    except (subprocess.CalledProcessError, json.JSONDecodeError, subprocess.TimeoutExpired):
        return None


def fetch_repositories():
    """Fetch all accessible repositories in parallel"""
    log("Fetching repositories in parallel...")

    if not CONFIG['personal_account']:
        user_info = run_gh_command(f"{CONFIG['gh_command']} api user --jq '.login'")
        if user_info:
            CONFIG['personal_account'] = str(user_info).strip('"')
            log(f"Detected personal account: {CONFIG['personal_account']}", verbose_only=True)

    if CONFIG['enable_rate_limit_check']:
        limit_data = check_rate_limit()
        if limit_data:
            log(f"📊 Initial Rate Limit: {limit_data['remaining']}/{limit_data['limit']} requests remaining")

    all_repos = []

    with ThreadPoolExecutor(max_workers=CONFIG['max_workers_fetch']) as executor:
        futures = []

        futures.append(executor.submit(
            run_gh_command,
            f"{CONFIG['gh_command']} repo list --json nameWithOwner,name,owner --limit 1000"
        ))

        orgs_future = executor.submit(
            run_gh_command,
            f"{CONFIG['gh_command']} api user/orgs --paginate"
        )
        orgs_data = orgs_future.result()

        if orgs_data:
            for org in orgs_data:
                futures.append(executor.submit(
                    run_gh_command,
                    f"{CONFIG['gh_command']} repo list {org['login']} --json nameWithOwner,name,owner --limit 1000"
                ))

        for future in as_completed(futures):
            result = future.result()
            if result:
                all_repos.extend(result)

    # Remove duplicates
    seen = set()
    unique_repos = []
    for repo in all_repos:
        if repo['nameWithOwner'] not in seen:
            seen.add(repo['nameWithOwner'])
            unique_repos.append(repo)

    return unique_repos


def decode_file_content(api_response):
    """Decode base64 file content from GitHub API response"""
    if not isinstance(api_response, dict):
        return None
    raw_content = api_response.get('content', '')
    if not raw_content:
        return None
    try:
        return base64.b64decode(raw_content).decode('utf-8')
    except Exception:
        return None


def get_custom_instruction_files(repo_name):
    """
    Retrieve all custom instruction files for a repository.

    Checks for:
    - .github/copilot-instructions.md
    - .github/instructions/*.instructions.md

    Returns a list of dicts with keys: path, char_count, exceeds_limit
    """
    files = []

    # Check root-level custom instructions file
    for file_path in CONFIG['custom_instruction_files']:
        api_path = f"repos/{repo_name}/contents/{file_path}"
        response = run_gh_command(f"{CONFIG['gh_command']} api {api_path}")
        content = decode_file_content(response)
        if content is not None:
            char_count = len(content)
            files.append({
                'path': file_path,
                'char_count': char_count,
                'exceeds_limit': char_count > CONFIG['max_chars'],
            })

    # Check .github/instructions/ directory for *.instructions.md files
    for dir_path in CONFIG['custom_instruction_dirs']:
        api_path = f"repos/{repo_name}/contents/{dir_path}"
        response = run_gh_command(f"{CONFIG['gh_command']} api {api_path}")
        if isinstance(response, list):
            for item in response:
                if (item.get('type') == 'file' and
                        item.get('name', '').endswith(CONFIG['custom_instruction_extension'])):
                    file_response = run_gh_command(
                        f"{CONFIG['gh_command']} api repos/{repo_name}/contents/{item['path']}"
                    )
                    content = decode_file_content(file_response)
                    if content is not None:
                        char_count = len(content)
                        files.append({
                            'path': item['path'],
                            'char_count': char_count,
                            'exceeds_limit': char_count > CONFIG['max_chars'],
                        })

    return files


def assess_repo(repo):
    """Assess a single repository for custom instruction file compliance"""
    full_name = repo['nameWithOwner']
    result = {
        'repo': full_name,
        'files_found': [],
        'violations': [],
        'has_custom_instructions': False,
        'error': None,
    }

    try:
        files = get_custom_instruction_files(full_name)
        result['files_found'] = files
        result['has_custom_instructions'] = len(files) > 0

        for f in files:
            if f['exceeds_limit']:
                result['violations'].append({
                    'repository': full_name,
                    'file_path': f['path'],
                    'char_count': f['char_count'],
                    'limit': CONFIG['max_chars'],
                    'excess_chars': f['char_count'] - CONFIG['max_chars'],
                })
    except Exception as e:
        result['error'] = str(e)

    return result


def check_all_repositories(repos):
    """Check all repositories for custom instruction file compliance in parallel"""
    log("\nValidating custom instruction files (parallel execution)...")

    results = []
    total = len(repos)

    with ThreadPoolExecutor(max_workers=CONFIG['max_workers_check']) as executor:
        future_to_repo = {executor.submit(assess_repo, repo): repo for repo in repos}

        completed = 0
        for future in as_completed(future_to_repo):
            result = future.result()
            results.append(result)
            completed += 1
            log(f"⚡ Progress: {completed}/{total} repositories checked ({(completed/total*100):.0f}%)", verbose_only=True)

    results.sort(key=lambda x: x['repo'])
    return results


def export_violations_to_excel(violations):
    """
    Export files exceeding the 4000-character limit to an Excel report.
    Returns the path of the created file, or None if no violations.
    """
    if not violations:
        return None

    if not OPENPYXL_AVAILABLE:
        log("⚠️  openpyxl not installed. Install with: pip install openpyxl")
        log("   Falling back to CSV output for violations.")
        return _export_violations_to_csv(violations)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{CONFIG['excel_prefix']}_{timestamp}.xlsx"
    filepath = Path(CONFIG['output_dir']) / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Repository",
        "File Path",
        "Character Count",
        f"Limit ({CONFIG['max_chars']} chars)",
        "Excess Characters",
        "GitHub Copilot Rule",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    rule_reference = (
        "Repository Custom Instructions must not exceed 4000 characters. "
        "See: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/"
        "response-customization?tool=webui#about-repository-custom-instructions"
    )

    for row_idx, violation in enumerate(violations, start=2):
        ws.cell(row=row_idx, column=1, value=violation['repository'])
        ws.cell(row=row_idx, column=2, value=violation['file_path'])
        ws.cell(row=row_idx, column=3, value=violation['char_count'])
        ws.cell(row=row_idx, column=4, value=violation['limit'])
        ws.cell(row=row_idx, column=5, value=violation['excess_chars'])
        ws.cell(row=row_idx, column=6, value=rule_reference)

    # Auto-fit column widths
    column_widths = [40, 50, 18, 18, 18, 80]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(filepath)
    log(f"✅ Excel violations report created: {filepath}")
    log(f"   Total violations: {len(violations)}")
    return str(filepath)


def _export_violations_to_csv(violations):
    """Fallback: export violations to CSV when openpyxl is not available"""
    import csv as csv_module

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{CONFIG['excel_prefix']}_{timestamp}.csv"
    filepath = Path(CONFIG['output_dir']) / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = ['repository', 'file_path', 'char_count', 'limit', 'excess_chars']
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv_module.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(violations)

    log(f"✅ CSV violations report created: {filepath}")
    log(f"   Total violations: {len(violations)}")
    return str(filepath)


def print_summary(results):
    """Print a summary of the assessment"""
    total_repos = len(results)
    repos_with_files = sum(1 for r in results if r['has_custom_instructions'])
    repos_with_violations = sum(1 for r in results if r['violations'])
    repos_with_errors = sum(1 for r in results if r['error'])
    total_files = sum(len(r['files_found']) for r in results)
    total_violations = sum(len(r['violations']) for r in results)

    log("\n" + "=" * 80)
    log("SUMMARY - CUSTOM INSTRUCTIONS VALIDATION")
    log("=" * 80)
    log(f"Total repositories scanned:            {total_repos}")
    log(f"Repositories with custom instructions: {repos_with_files}")
    log(f"Total custom instruction files found:  {total_files}")
    log(f"Repositories with violations:          {repos_with_violations}")
    log(f"Total files exceeding {CONFIG['max_chars']} chars:       {total_violations}")
    log(f"Repositories with errors:              {repos_with_errors}")

    if repos_with_violations > 0:
        log(f"\n❌ VIOLATIONS (files exceeding {CONFIG['max_chars']} characters):")
        for result in results:
            for v in result['violations']:
                log(f"   • {v['repository']} | {v['file_path']} | {v['char_count']} chars (+{v['excess_chars']} over limit)")
        log(f"\n📖 GitHub Copilot Rule: Repository Custom Instructions must not exceed {CONFIG['max_chars']} characters.")
        log("   Reference: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/response-customization?tool=webui#about-repository-custom-instructions")


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_existence(results):
    """
    Validate that at least one custom instructions file exists across all repositories.
    Returns True if at least one file is found, False otherwise.
    """
    total_files = sum(len(r['files_found']) for r in results)
    if total_files == 0:
        log("\n❌ VALIDATION FAILED: No Repository Custom Instructions files found.")
        log("   GitHub Copilot supports Repository Custom Instructions configured via:")
        log("     - .github/copilot-instructions.md")
        log("     - .github/instructions/*.instructions.md")
        log("   Reference: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/response-customization?tool=webui#about-repository-custom-instructions")
        return False
    log(f"\n✅ VALIDATION PASSED: Found {total_files} custom instruction file(s).")
    return True


def validate_file_sizes(results):
    """
    Validate that all custom instruction files are within the 4000-character limit.
    Returns list of violations (empty list means all files are valid).
    """
    all_violations = []
    for result in results:
        all_violations.extend(result['violations'])

    if all_violations:
        log(f"\n❌ VALIDATION FAILED: {len(all_violations)} file(s) exceed the {CONFIG['max_chars']}-character limit.")
        log(f"   GitHub Copilot Rule: Repository Custom Instructions must not exceed {CONFIG['max_chars']} characters.")
        log("   Reference: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/response-customization?tool=webui#about-repository-custom-instructions")
    else:
        log(f"\n✅ VALIDATION PASSED: All custom instruction files are within the {CONFIG['max_chars']}-character limit.")

    return all_violations


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function"""
    start_time = time()

    log("=" * 80)
    log("GITHUB COPILOT CUSTOM INSTRUCTIONS VALIDATION TOOL")
    log("=" * 80)
    log(f"Validation Rules:")
    log(f"  • Existence: At least one custom instruction file must be present")
    log(f"  • Size Limit: Each file must not exceed {CONFIG['max_chars']} characters")
    log(f"  • Reference: https://docs.github.com/en/enterprise-cloud@latest/copilot/concepts/prompting/response-customization")

    if not check_gh_installed():
        log("\n❌ GitHub CLI (gh) is not installed or not in PATH!")
        log("\nTo install:")
        log("  Windows: winget install --id GitHub.cli")
        log("  macOS:   brew install gh")
        log("  Linux:   See https://cli.github.com/")
        log("\nAfter installation, authenticate with: gh auth login")
        return 1

    # Fetch repositories
    fetch_start = time()
    repos = fetch_repositories()
    fetch_time = time() - fetch_start

    if not repos:
        log("\n❌ Could not fetch repositories. Make sure you're authenticated:")
        log("  gh auth login")
        return 1

    log(f"✓ Found {len(repos)} repositories in {fetch_time:.2f}s")

    # Check all repositories
    check_start = time()
    results = check_all_repositories(repos)
    check_time = time() - check_start

    # Print summary
    print_summary(results)

    log(f"\n⚡ PERFORMANCE METRICS:")
    log(f"   Repository fetch: {fetch_time:.2f}s")
    log(f"   Validation check: {check_time:.2f}s")
    log(f"   Total execution:  {(time() - start_time):.2f}s")

    # --- Validation 1: Existence ---
    existence_ok = validate_existence(results)
    if not existence_ok:
        log("\n" + "=" * 80)
        log("❌ Validation failed: no custom instruction files found.")
        log("=" * 80)
        return 1

    # --- Validation 2: File size ---
    violations = validate_file_sizes(results)
    if violations:
        export_violations_to_excel(violations)
        log("\n" + "=" * 80)
        log("❌ Validation failed: one or more custom instruction files exceed the 4000-character limit.")
        log("=" * 80)
        return 1

    log("\n" + "=" * 80)
    log("✅ All custom instruction validations passed!")
    log("=" * 80)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        log("\n\n⚠️  Operation cancelled by user")
        sys.exit(130)
    except Exception as e:
        log(f"\n❌ Unexpected error: {e}")
        sys.exit(1)
